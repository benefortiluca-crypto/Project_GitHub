import os
import sys
import json
import datetime
import traceback
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from typing import List, Dict, Any, Optional

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import pandas as pd

import db

# Initialize FastAPI app
app = FastAPI(title="VDL Manager Pro", description="Gestione Vendor Document List & Invio Reminder")

# Add CORS Middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure error logging file in the project folder
ERROR_LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backend_errors.log")

def log_error_to_file(message: str, exc: Exception = None):
    try:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(ERROR_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] ERROR: {message}\n")
            if exc:
                f.write("Traceback:\n")
                f.write("".join(traceback.format_exception(type(exc), exc, exc.__traceback__)))
            f.write("="*80 + "\n")
    except Exception as e:
        print(f"Failed to write to error log: {e}")

from fastapi import Request
@app.middleware("http")
async def error_logging_middleware(request: Request, call_next):
    try:
        response = await call_next(request)
        return response
    except Exception as exc:
        log_error_to_file(f"Unhandled exception during {request.method} {request.url.path}", exc)
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": f"Errore interno del server: {str(exc)}"}
        )


# Standard lifecycle tracking columns to append
CUSTOM_TRACKING_COLUMNS = [
    "Internal Document Number",
    "Hands",
    "1° Invio previsione",
    "Promise date",
    "Next issue forecast date",
    "Last Code receive",
    "TR Out",
    "Actual Date",
    "TR In",
    "Return Code",
    "Return Date",
    "TR Out1",
    "Actual Date1",
    "TR In1",
    "Return Code1",
    "Return Date1",
    "TR Out2",
    "Actual Date2",
    "TR In2",
    "Return Code2",
    "Return Date2",
    "TR Out3",
    "Actual Date3",
    "TR In3",
    "Return Code3",
    "Return Date3",
    "TR Out4",
    "Actual Date4",
    "TR In4",
    "Return Code4",
    "Return Date4",
    "TR Out5",
    "Actual Date5",
    "TR In5",
    "Return Code5",
    "Return Date5",
    "TR Out6",
    "Actual Date6",
    "TR In6",
    "Return Code6",
    "Return Date6"
]

# Order for scanning dates from right (latest stage) to left (earliest stage)
SCAN_DATE_COLUMNS_ORDER = [
    "Return Date6", "Actual Date6",
    "Return Date5", "Actual Date5",
    "Return Date4", "Actual Date4",
    "Return Date3", "Actual Date3",
    "Return Date2", "Actual Date2",
    "Return Date1", "Actual Date1",
    "Return Date", "Actual Date",
    "Promise date", "1° Invio previsione"
]

# Ensure DB is initialized
db.init_db()

# DIAGNOSTIC CODE
try:
    import sqlite3
    diag_conn = sqlite3.connect(db.DB_PATH)
    diag_conn.row_factory = sqlite3.Row
    diag_cursor = diag_conn.cursor()
    
    diag_info = []
    diag_info.append("=== DB DIAGNOSTIC ===")
    diag_info.append(f"DB Path: {db.DB_PATH}")
    diag_info.append(f"Exists: {os.path.exists(db.DB_PATH)}")
    if os.path.exists(db.DB_PATH):
        diag_info.append(f"Size: {os.path.getsize(db.DB_PATH)} bytes")
        
        # Get tables
        tables = diag_cursor.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
        diag_info.append(f"Tables: {[t['name'] for t in tables]}")
        
        for t in tables:
            t_name = t['name']
            count = diag_cursor.execute(f"SELECT COUNT(*) FROM {t_name}").fetchone()[0]
            diag_info.append(f"Table '{t_name}' row count: {count}")
            
        # Get users
        users = diag_cursor.execute("SELECT id, username, role, password, created_at FROM users").fetchall()
        diag_info.append("Users:")
        for u in users:
            diag_info.append(f"  - ID: {u['id']}, Username: {u['username']}, Role: {u['role']}, Password: {u['password']}, CreatedAt: {u['created_at']}")
            
        # Get projects
        projects = diag_cursor.execute("SELECT id, project_name FROM vdl_projects").fetchall()
        diag_info.append("Projects:")
        for p in projects:
            diag_info.append(f"  - ID: {p['id']}, Name: {p['project_name']}")
            
    diag_conn.close()
except Exception as diag_ex:
    diag_info = [f"DIAGNOSTIC ERROR: {str(diag_ex)}", traceback.format_exc()]

with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "db_diagnostic.txt"), "w", encoding="utf-8") as diag_f:
    diag_f.write("\n".join(diag_info))

# Models
class ProjectCreate(BaseModel):
    company: str
    contractor: str
    contractor_proj_num: str
    vendor_proj_num: str
    project_name: str

class DocumentUpdate(BaseModel):
    document_data: Dict[str, Any]

class DocumentAdd(BaseModel):
    project_id: int
    document_data: Dict[str, Any]

class ContactSave(BaseModel):
    project_id: int
    hands_value: str
    to_emails: str
    cc_emails: str

class ProjectExportRequest(BaseModel):
    project_id: int
    columns: List[str]
    documents: List[Dict[str, Any]]

class SettingsSave(BaseModel):
    sender_email: str
    smtp_server: str
    smtp_port: str
    smtp_user: str
    smtp_password: str
    email_mode: str
    exchange_time: str

class ReminderCalculateRequest(BaseModel):
    project_id: int
    hands_value: str
    exchange_time: int
    language: str = "it"

class EmailSendRequest(BaseModel):
    project_id: int
    hands_value: str
    exchange_time: int
    sender_email: str
    subject: str
    additional_notes: Optional[str] = ""
    language: str = "it"

# Auth Models
class UserLogin(BaseModel):
    username: str
    password: str

class UserRegister(BaseModel):
    username: str
    password: str
    role: str

class UserAdminCreate(BaseModel):
    username: str
    password: str
    role: str

class UserAdminUpdate(BaseModel):
    username: str
    password: str
    role: str

# New Pydantic models for Suppliers and Supplier Transmittals
class SupplierCreate(BaseModel):
    name: str
    item: str

class SupplierUpdate(BaseModel):
    name: str
    item: str

class ProjectSuppliersSync(BaseModel):
    supplier_ids: List[int]

class SupplierTransmittalCreate(BaseModel):
    project_id: int
    supplier_id: int
    direction: str  # 'IN' or 'OUT'
    tr_number: str
    tr_date: str
    notes: Optional[str] = ""
    document_list: List[Dict[str, Any]]


# Utility function to parse date robustly
def parse_date(val: Any) -> Optional[datetime.date]:
    if val is None or pd.isna(val):
        return None
    
    if isinstance(val, (datetime.datetime, datetime.date)):
        if isinstance(val, datetime.datetime):
            return val.date()
        return val
        
    # Convert to string and strip whitespace
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ["nan", "nat", "null", "-", ""]:
        return None
        
    # Try different string formats
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%Y/%m/%d"
    ]
    
    for fmt in formats:
        try:
            dt = datetime.datetime.strptime(val_str, fmt)
            return dt.date()
        except ValueError:
            continue
            
    # Try Excel numeric float parsing
    try:
        # Excel epoch starts 1900-01-01, but has a leap year bug in 1900
        excel_num = float(val_str)
        if excel_num > 60:
            excel_num -= 1
        dt = datetime.date(1899, 12, 30) + datetime.timedelta(days=excel_num)
        return dt
    except ValueError:
        pass
        
    return None

def get_row_val(row: Dict[str, Any], field_name: str) -> Any:
    norm_field = field_name.lower().replace(" ", "").strip()
    for k, v in row.items():
        if k.lower().replace(" ", "").strip() == norm_field:
            return v
    return None

def get_doc_title(row: Dict[str, Any]) -> str:
    # 1. Fuzzy match for Title/Description
    for k, v in row.items():
        if k == "__id":
            continue
        k_lower = k.lower().replace(" ", "").replace("_", "").replace("-", "").strip()
        if any(term in k_lower for term in ["title", "description", "titolo", "descrizione", "nomedocumento", "documentname"]):
            if v and str(v).strip():
                return str(v).strip()
                
    # 2. Fallback check for Name/Nome/Documento
    for k, v in row.items():
        if k == "__id":
            continue
        k_lower = k.lower().replace(" ", "").replace("_", "").replace("-", "").strip()
        if any(term in k_lower for term in ["name", "nome", "documento"]):
            if v and str(v).strip():
                return str(v).strip()
                
    return "Nessuna descrizione"

def get_doc_code(row: Dict[str, Any]) -> str:
    # 1. Fuzzy match for document code / number keys
    for k, v in row.items():
        if k == "__id":
            continue
        k_lower = k.lower().replace(" ", "").replace("_", "").replace("-", "").strip()
        if ("doc" in k_lower and any(term in k_lower for term in ["code", "number", "no", "num", "codice"])) or k_lower in ["internaldocumentnumber", "documentnumber", "codicedocumento", "codicedoc"]:
            if v and str(v).strip():
                return str(v).strip()
    # 2. Fallback exact matches
    v1 = get_row_val(row, "Internal Document Number")
    if v1: return str(v1).strip()
    v2 = get_row_val(row, "Document Number")
    if v2: return str(v2).strip()
    v3 = get_row_val(row, "Codice Documento")
    if v3: return str(v3).strip()
    
    # 3. First non-empty key
    for k, v in row.items():
        if k != "__id" and v and str(v).strip():
            return str(v).strip()
    return "-"


def set_row_val(row: Dict[str, Any], field_name: str, value: Any):
    norm_field = field_name.lower().replace(" ", "").strip()
    for k in row.keys():
        if k.lower().replace(" ", "").strip() == norm_field:
            row[k] = value
            return
    # If not found, add it with the original field_name
    row[field_name] = value

def recalculate_computed_fields(row: Dict[str, Any], exchange_time_days: int, project_id: Optional[int] = None) -> Dict[str, Any]:
    import re
    # Find all Return Code and Return Date columns dynamically
    code_cols_map = {}
    date_cols_map = {}
    
    for k in row.keys():
        norm_k = k.lower().replace(" ", "").strip()
        if norm_k.startswith("returncode"):
            match = re.search(r'returncode(\d*)', norm_k)
            if match:
                num_str = match.group(1)
                num = int(num_str) if num_str else 0
                code_cols_map[num] = k
        elif norm_k.startswith("returndate"):
            match = re.search(r'returndate(\d*)', norm_k)
            if match:
                num_str = match.group(1)
                num = int(num_str) if num_str else 0
                date_cols_map[num] = k
                
    # Sort descending by cycle number (latest first)
    sorted_code_nums = sorted(code_cols_map.keys(), reverse=True)
    sorted_date_nums = sorted(date_cols_map.keys(), reverse=True)
    
    # 1. Calculate Last Code receive
    last_code = ""
    for num in sorted_code_nums:
        col = code_cols_map[num]
        val = get_row_val(row, col)
        if val is not None:
            val_str = str(val).strip()
            if val_str and val_str.lower() not in ["nan", "nat", "null", "-"]:
                last_code = val_str
                break
    set_row_val(row, "Last Code receive", last_code)

    # 2. Calculate Next issue forecast date
    next_forecast_str = ""
    last_return_date = None
    for num in sorted_date_nums:
        col = date_cols_map[num]
        val = get_row_val(row, col)
        if val:
            dt = parse_date(val)
            if dt:
                last_return_date = dt
                break
                
    if last_return_date:
        next_forecast = last_return_date + datetime.timedelta(days=exchange_time_days)
        next_forecast_str = next_forecast.strftime("%Y-%m-%d")
    else:
        # Fallback to Promise date
        promise_val = get_row_val(row, "Promise date")
        promise_dt = parse_date(promise_val) if promise_val else None
        if promise_dt:
            next_forecast_str = promise_dt.strftime("%Y-%m-%d")
        else:
            # Fallback to 1° Invio previsione
            prev_val = get_row_val(row, "1° Invio previsione")
            prev_dt = parse_date(prev_val) if prev_val else None
            if prev_dt:
                next_forecast_str = prev_dt.strftime("%Y-%m-%d")
                
    set_row_val(row, "Next issue forecast date", next_forecast_str)

    # 3. Calculate Hands automation rule
    if project_id:
        project = db.get_project(project_id)
        if project:
            contractor = project.get("contractor") or ""
            # Initialize _original_hands if not set
            hands_val = get_row_val(row, "Hands")
            orig_hands = row.get("_original_hands") or row.get("original_hands")
            if not orig_hands:
                if hands_val:
                    row["_original_hands"] = hands_val
                    orig_hands = hands_val
                else:
                    row["_original_hands"] = ""
                    orig_hands = ""
            
            # Find the latest cycle with a filled TR Out
            latest_cycle_index = -1
            latest_cycle_has_tr_in = False
            
            # Cycle 0
            tr_out_0 = get_row_val(row, "TR Out")
            tr_in_0 = get_row_val(row, "TR In")
            if tr_out_0 and str(tr_out_0).strip() and str(tr_out_0).strip() not in ["-", "nan", "nat", "null", ""]:
                latest_cycle_index = 0
                latest_cycle_has_tr_in = bool(tr_in_0 and str(tr_in_0).strip() and str(tr_in_0).strip() not in ["-", "nan", "nat", "null", ""])
                
            # Cycles 1 to 6
            for idx in range(1, 7):
                tr_out_col = f"TR Out{idx}"
                tr_in_col = f"TR In{idx}"
                tr_out_val = get_row_val(row, tr_out_col)
                tr_in_val = get_row_val(row, tr_in_col)
                if tr_out_val and str(tr_out_val).strip() and str(tr_out_val).strip() not in ["-", "nan", "nat", "null", ""]:
                    latest_cycle_index = idx
                    latest_cycle_has_tr_in = bool(tr_in_val and str(tr_in_val).strip() and str(tr_in_val).strip() not in ["-", "nan", "nat", "null", ""])
            
            if latest_cycle_index != -1:
                if latest_cycle_has_tr_in:
                    set_row_val(row, "Hands", contractor)
                else:
                    set_row_val(row, "Hands", orig_hands)
            else:
                set_row_val(row, "Hands", orig_hands)

    return row

def recalculate_row_list(rows: List[Dict[str, Any]], project_id: Optional[int] = None) -> List[Dict[str, Any]]:
    settings = db.get_settings()
    exchange_time = int(settings.get("exchange_time", "15") or "15")
    for row in rows:
        recalculate_computed_fields(row, exchange_time, project_id)
    return rows


# Security dependencies
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import uuid

security = HTTPBearer(auto_error=False)

def get_current_user(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)):
    # Bypassed authentication to disable access modes
    return {
        "id": 1,
        "username": "ADMIN",
        "role": "Admin",
        "created_at": "2026-06-16 00:00:00"
    }

def require_roles(allowed_roles: List[str]):
    def dependency(current_user: dict = Depends(get_current_user)):
        if current_user["role"] not in allowed_roles:
            raise HTTPException(
                status_code=403, 
                detail=f"Permesso negato: l'azione richiede uno dei seguenti ruoli: {', '.join(allowed_roles)}."
            )
        return current_user
    return dependency


# ==========================================
# AUTH ENDPOINTS
# ==========================================

@app.post("/api/auth/login")
def login(req: UserLogin):
    user = db.get_user_by_username(req.username)
    if not user or user["password"] != req.password:
        raise HTTPException(status_code=400, detail="Credenziali non valide. Riprova.")
        
    token = str(uuid.uuid4())
    db.create_session(user["id"], token)
    return {
        "status": "success",
        "token": token,
        "username": user["username"],
        "role": user["role"]
    }

@app.post("/api/auth/register")
def register(req: UserRegister):
    if req.role not in ["Project Manager", "Project Engineering", "Document Controller"]:
        raise HTTPException(status_code=400, detail="Ruolo non valido.")
        
    user_id = db.create_user(req.username, req.password, req.role)
    if not user_id:
        raise HTTPException(status_code=400, detail="Username già esistente.")
        
    return {"status": "success", "message": "Registrazione completata. Effettua il login."}

@app.post("/api/auth/logout")
def logout(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)):
    if credentials:
        db.delete_session(credentials.credentials)
    return {"status": "success"}

@app.get("/api/auth/me")
def get_me(current_user: dict = Depends(get_current_user)):
    return {"status": "success", "user": {
        "username": current_user["username"],
        "role": current_user["role"]
    }}

# ==========================================
# ADMIN USER MANAGEMENT ENDPOINTS
# ==========================================

@app.get("/api/admin/users")
def list_users(current_user: dict = Depends(require_roles(["Admin"]))):
    users = db.get_all_users()
    return {"status": "success", "users": users}

@app.post("/api/admin/users")
def admin_create_user(req: UserAdminCreate, current_user: dict = Depends(require_roles(["Admin"]))):
    if req.role not in ["Admin", "Project Manager", "Project Engineering", "Document Controller"]:
        raise HTTPException(status_code=400, detail="Ruolo non valido.")
        
    user_id = db.create_user(req.username, req.password, req.role)
    if not user_id:
        raise HTTPException(status_code=400, detail="Username già esistente.")
        
    return {"status": "success", "message": "Utente creato con successo."}

@app.put("/api/admin/users/{user_id}")
def admin_update_user(user_id: int, req: UserAdminUpdate, current_user: dict = Depends(require_roles(["Admin"]))):
    if req.role not in ["Admin", "Project Manager", "Project Engineering", "Document Controller"]:
        raise HTTPException(status_code=400, detail="Ruolo non valido.")
        
    db.update_user(user_id, req.username, req.password, req.role)
    return {"status": "success", "message": "Utente aggiornato con successo."}

@app.delete("/api/admin/users/{user_id}")
def admin_delete_user(user_id: int, current_user: dict = Depends(require_roles(["Admin"]))):
    if current_user["id"] == user_id:
        raise HTTPException(status_code=400, detail="Non puoi eliminare la tua stessa utenza amministratore.")
        
    db.delete_user(user_id)
    return {"status": "success", "message": "Utente eliminato con successo."}


# Endpoints
@app.post("/api/project/new")
def create_new_project(req: ProjectCreate, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    pid = db.create_project(req.company, req.contractor, req.contractor_proj_num, req.vendor_proj_num, req.project_name)
    return {"status": "success", "project_id": pid}

@app.put("/api/project/{project_id}")
def update_project(project_id: int, req: ProjectCreate, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Progetto non trovato.")
    db.update_project_metadata(project_id, req.company, req.contractor, req.contractor_proj_num, req.vendor_proj_num, req.project_name)
    return {"status": "success", "message": "Progetto aggiornato con successo."}

@app.get("/api/projects")
def get_projects(current_user: dict = Depends(get_current_user)):
    projects = db.get_all_projects()
    return {"status": "success", "projects": projects}

@app.get("/api/project/{project_id}")
def get_project(project_id: int, current_user: dict = Depends(get_current_user)):
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Progetto non trovato.")
    return {"status": "success", "project": project}

@app.post("/api/upload")
async def upload_vdl(project_id: int = Form(...), file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    try:
        filename = file.filename
        content = await file.read()
        
        # Determine format
        if filename.endswith(".csv"):
            import io
            df = pd.read_csv(io.BytesIO(content))
        elif filename.endswith((".xlsx", ".xls")):
            import io
            df = pd.read_excel(io.BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Formato file non supportato. Carica un file Excel o CSV.")
            
        # Convert datetime columns to strings
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime('%Y-%m-%d')
                
        # Clean data safely: replace NaNs and handle edge case datetimes
        raw_rows = df.to_dict(orient="records")
        cleaned_rows = []
        import datetime
        for row in raw_rows:
            clean_row = {}
            for k, v in row.items():
                if pd.isna(v):
                    clean_row[k] = ""
                elif isinstance(v, (datetime.datetime, datetime.date)):
                    clean_row[k] = v.strftime('%Y-%m-%d')
                else:
                    clean_row[k] = v
            cleaned_rows.append(clean_row)
            
        df = pd.DataFrame(cleaned_rows)
        
        # Check current columns and append missing custom columns
        existing_cols = list(df.columns)
        
        # Normalize and find which custom columns are missing
        columns_to_add = []
        for custom_col in CUSTOM_TRACKING_COLUMNS:
            # Let's do a case-insensitive, space-flexible check
            normalized_custom = custom_col.lower().replace(" ", "").strip()
            found = False
            for exist_col in existing_cols:
                if exist_col.lower().replace(" ", "").strip() == normalized_custom:
                    found = True
                    break
            if not found:
                columns_to_add.append(custom_col)
                
        # Append missing columns to DataFrame with empty values
        for col in columns_to_add:
            df[col] = None
            
        final_columns = list(df.columns)
        rows = df.to_dict(orient="records")
        
        # Recalculate computed fields for all uploaded rows!
        settings = db.get_settings()
        exchange_time = int(settings.get("exchange_time", "15") or "15")
        for row in rows:
            recalculate_computed_fields(row, exchange_time)
            
        # Save in SQLite
        db.save_vdl_project(project_id, final_columns, rows)
        
        return {
            "status": "success",
            "message": f"VDL caricata con successo con {len(rows)} righe nel progetto.",
            "project_id": project_id
        }
    except Exception as e:
        log_error_to_file("Errore durante il caricamento della VDL", e)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Errore durante il caricamento del file: {str(e)}")

@app.get("/api/admin/error-log")
def download_error_log(current_user: dict = Depends(get_current_user)):
    if not os.path.exists(ERROR_LOG_PATH):
        try:
            with open(ERROR_LOG_PATH, "w", encoding="utf-8") as f:
                f.write("=== Backend Error Log Initialized ===\n")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Impossibile creare il file di log: {str(e)}")
    return FileResponse(ERROR_LOG_PATH, filename="backend_errors.log", media_type="text/plain")

@app.post("/api/admin/error-log/clear")
def clear_error_log(current_user: dict = Depends(get_current_user)):
    try:
        with open(ERROR_LOG_PATH, "w", encoding="utf-8") as f:
            f.write(f"=== Backend Error Log Cleared at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n")
        return {"status": "success", "message": "Log cancellato con successo."}
    except Exception as e:
        log_error_to_file("Errore durante la cancellazione del log", e)
        raise HTTPException(status_code=500, detail=f"Errore: {str(e)}")


@app.get("/api/documents")
def get_documents(project_id: int, current_user: dict = Depends(get_current_user)):
    docs = db.get_documents(project_id)
    return {"status": "success", "documents": recalculate_row_list(docs, project_id)}

@app.post("/api/documents")
def add_new_document(doc: DocumentAdd, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    project = db.get_project(doc.project_id)
    if not project:
        raise HTTPException(status_code=400, detail="Progetto non trovato.")
    
    # Fill in all project columns with blank if not present
    full_data = {}
    for col in project["columns"]:
        full_data[col] = doc.document_data.get(col, "")
        
    settings = db.get_settings()
    exchange_time = int(settings.get("exchange_time", "15") or "15")
    recalculate_computed_fields(full_data, exchange_time, doc.project_id)
        
    doc_id = db.add_document(doc.project_id, full_data)
    return {"status": "success", "message": "Documento aggiunto.", "id": doc_id}

@app.put("/api/documents/{doc_id}")
def update_existing_document(doc_id: int, doc: DocumentUpdate, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    project_id = db.get_project_id_by_doc_id(doc_id)
    settings = db.get_settings()
    exchange_time = int(settings.get("exchange_time", "15") or "15")
    recalculate_computed_fields(doc.document_data, exchange_time, project_id)
    db.update_document(doc_id, doc.document_data)
    return {"status": "success", "message": "Documento aggiornato."}

@app.delete("/api/documents/{doc_id}")
def delete_existing_document(doc_id: int, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    db.delete_document(doc_id)
    return {"status": "success", "message": "Documento eliminato."}

@app.post("/api/documents/save-all")
def save_all_documents(project_id: int, payload: List[Dict[str, Any]], current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    settings = db.get_settings()
    exchange_time = int(settings.get("exchange_time", "15") or "15")
    # This bulk saves modifications from the table
    for row in payload:
        if "__id" in row:
            doc_id = row["__id"]
            # Strip out temporary metadata
            cleaned_row = {k: v for k, v in row.items() if k != "__id" and k != "is_dirty"}
            recalculate_computed_fields(cleaned_row, exchange_time, project_id)
            db.update_document(doc_id, cleaned_row)
            
    return {"status": "success", "message": "Tutte le modifiche sono state salvate nel DB."}

# ==========================================
# PROJECT SPECIFIC SETTINGS & FILES
# ==========================================

class ProjectSettingsSave(BaseModel):
    job_path: str
    revision_format: str
    revision_columns: List[str]

class OpenFileRequest(BaseModel):
    filepath: str
    show_in_folder: bool = False

def is_safe_path(base_dir: str, path: str, follow_symlinks: bool = True) -> bool:
    if not base_dir:
        return False
    if follow_symlinks:
        matchpath = os.path.realpath(path)
        directory = os.path.realpath(base_dir)
    else:
        matchpath = os.path.abspath(path)
        directory = os.path.abspath(base_dir)
    return matchpath.startswith(directory + os.sep) or matchpath == directory

import re
def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', '_', name).strip()

@app.post("/api/project/{project_id}/settings")
def save_project_settings(project_id: int, req: ProjectSettingsSave, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller", "Project Engineering"]))):
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Progetto non trovato.")
    
    db.update_project_path_and_revisions(project_id, req.job_path, req.revision_format, req.revision_columns)
    
    if req.job_path and req.job_path.strip():
        base_dir = req.job_path.strip()
        try:
            os.makedirs(os.path.join(base_dir, "Customer Specification"), exist_ok=True)
            os.makedirs(os.path.join(base_dir, "Job Documentation"), exist_ok=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Impossibile creare le cartelle nel percorso specificato: {str(e)}")
            
    return {"status": "success", "message": "Impostazioni di progetto salvate con successo."}

@app.get("/api/project/{project_id}/specifications")
def list_specifications(project_id: int, current_user: dict = Depends(get_current_user)):
    project = db.get_project(project_id)
    if not project or not project.get("job_path"):
        return {"status": "success", "files": []}
        
    spec_dir = os.path.join(project["job_path"], "Customer Specification")
    if not os.path.exists(spec_dir):
        return {"status": "success", "files": []}
        
    files_list = []
    try:
        for f in os.listdir(spec_dir):
            full_path = os.path.join(spec_dir, f)
            if os.path.isfile(full_path):
                stat = os.stat(full_path)
                mtime = datetime.datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                _, ext = os.path.splitext(f)
                files_list.append({
                    "name": f,
                    "type": ext.upper().replace(".", "") or "FILE",
                    "size": stat.st_size,
                    "modified": mtime,
                    "path": full_path
                })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore nella lettura delle specifiche: {str(e)}")
        
    return {"status": "success", "files": files_list}

@app.post("/api/project/{project_id}/specifications/upload")
async def upload_specification(project_id: int, file: UploadFile = File(...), current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller", "Project Engineering"]))):
    project = db.get_project(project_id)
    if not project or not project.get("job_path"):
        raise HTTPException(status_code=400, detail="Configura prima il percorso JOB nelle impostazioni.")
        
    spec_dir = os.path.join(project["job_path"], "Customer Specification")
    os.makedirs(spec_dir, exist_ok=True)
    
    filename = sanitize_filename(file.filename)
    dest_path = os.path.join(spec_dir, filename)
    try:
        content = await file.read()
        with open(dest_path, "wb") as f:
            f.write(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore nel salvataggio del file: {str(e)}")
        
    return {"status": "success", "message": "Specifica caricata con successo."}

@app.delete("/api/project/{project_id}/specifications")
def delete_specification(project_id: int, filename: str, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller", "Project Engineering"]))):
    project = db.get_project(project_id)
    if not project or not project.get("job_path"):
        raise HTTPException(status_code=400, detail="Percorso di progetto non configurato.")
        
    spec_dir = os.path.join(project["job_path"], "Customer Specification")
    file_path = os.path.join(spec_dir, filename)
    
    if not is_safe_path(spec_dir, file_path):
        raise HTTPException(status_code=400, detail="Accesso al file non consentito.")
        
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Errore nella cancellazione: {str(e)}")
            
    return {"status": "success", "message": "File eliminato."}

@app.get("/api/project/{project_id}/specifications/search")
def search_specifications(project_id: int, query: str, current_user: dict = Depends(get_current_user)):
    if not query or not query.strip():
        return {"status": "success", "results": []}
        
    project = db.get_project(project_id)
    if not project or not project.get("job_path"):
        raise HTTPException(status_code=400, detail="Configura prima il percorso JOB nelle impostazioni.")
        
    spec_dir = os.path.join(project["job_path"], "Customer Specification")
    if not os.path.exists(spec_dir):
        return {"status": "success", "results": []}
        
    # Attempt to import fitz, installing it dynamically if missing
    try:
        import fitz
    except ImportError:
        import subprocess
        import sys
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "pymupdf"])
            import fitz
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Libreria PyMuPDF (fitz) mancante e installazione fallita: {str(e)}")
            
    if not fitz:
        raise HTTPException(status_code=500, detail="Libreria PyMuPDF (fitz) non disponibile per l'estrazione di testi dai PDF.")
        
    query_clean = query.lower().strip()
    results = []
    
    try:
        for f in os.listdir(spec_dir):
            filepath = os.path.join(spec_dir, f)
            if not os.path.isfile(filepath) or not f.lower().endswith(".pdf"):
                continue
                
            try:
                doc = fitz.open(filepath)
                for page_num in range(len(doc)):
                    page = doc[page_num]
                    text = page.get_text("text")
                    if not text:
                        continue
                        
                    text_lower = text.lower()
                    start_idx = 0
                    while True:
                        idx = text_lower.find(query_clean, start_idx)
                        if idx == -1:
                            break
                            
                        snippet_start = max(0, idx - 120)
                        snippet_end = min(len(text), idx + len(query_clean) + 120)
                        snippet = text[snippet_start:snippet_end].replace("\n", " ")
                        
                        if snippet_start > 0:
                            snippet = "..." + snippet
                        if snippet_end < len(text):
                            snippet = snippet + "..."
                            
                        results.append({
                            "filename": f,
                            "page": page_num + 1,
                            "snippet": snippet,
                            "filepath": filepath
                        })
                        
                        if len(results) >= 50:
                            break
                            
                        start_idx = idx + len(query_clean)
                        
                    if len(results) >= 50:
                        break
            except Exception as pe:
                print(f"Error reading PDF {f}: {str(pe)}")
                continue
                
            if len(results) >= 50:
                break
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore durante la ricerca nelle specifiche: {str(e)}")
        
    return {"status": "success", "results": results[:50]}

@app.post("/api/project/{project_id}/open-file")
def open_file_locally(project_id: int, req: OpenFileRequest, current_user: dict = Depends(get_current_user)):
    project = db.get_project(project_id)
    if not project or not project.get("job_path"):
        raise HTTPException(status_code=400, detail="Percorso JOB non configurato per il progetto.")
        
    if not is_safe_path(project["job_path"], req.filepath):
        raise HTTPException(status_code=403, detail="Non autorizzato ad accedere a file al di fuori della cartella di progetto.")
        
    if not os.path.exists(req.filepath):
        raise HTTPException(status_code=404, detail="File non trovato.")
        
    try:
        import subprocess
        if req.show_in_folder:
            win_path = os.path.normpath(req.filepath)
            subprocess.Popen(f'explorer.exe /select,"{win_path}"')
        else:
            os.startfile(req.filepath)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Impossibile aprire il file: {str(e)}")
        
    return {"status": "success", "message": "File aperto."}

@app.get("/api/project/{project_id}/vdl-documents/files")
def list_linked_vdl_documents(project_id: int, current_user: dict = Depends(get_current_user)):
    project = db.get_project(project_id)
    if not project or not project.get("job_path"):
        return {"status": "success", "files": {}}
        
    doc_dir = os.path.join(project["job_path"], "Job Documentation")
    if not os.path.exists(doc_dir):
        return {"status": "success", "files": {}}
        
    mapping = {}
    try:
        for f in os.listdir(doc_dir):
            full_path = os.path.join(doc_dir, f)
            if os.path.isfile(full_path):
                stem, ext = os.path.splitext(f)
                mapping[stem] = {
                    "filename": f,
                    "filepath": full_path
                }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore nella lettura dei file documentazione: {str(e)}")
        
    return {"status": "success", "files": mapping}

@app.post("/api/project/{project_id}/vdl-document/upload")
async def upload_vdl_document(project_id: int, document_number: str = Form(...), file: UploadFile = File(...), current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller", "Project Engineering"]))):
    project = db.get_project(project_id)
    if not project or not project.get("job_path"):
        raise HTTPException(status_code=400, detail="Configura prima il percorso JOB nelle impostazioni del progetto.")
        
    doc_dir = os.path.join(project["job_path"], "Job Documentation")
    os.makedirs(doc_dir, exist_ok=True)
    
    sanitized_num = sanitize_filename(document_number)
    _, ext = os.path.splitext(file.filename)
    dest_name = f"{sanitized_num}{ext}"
    dest_path = os.path.join(doc_dir, dest_name)
    
    try:
        content = await file.read()
        with open(dest_path, "wb") as f:
            f.write(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore nel salvataggio della documentazione: {str(e)}")
        
    return {"status": "success", "message": "Documento caricato con successo.", "filename": dest_name, "filepath": dest_path}

@app.post("/api/project/export")
def export_project_vdl(req: ProjectExportRequest, current_user: dict = Depends(get_current_user)):
    try:
        import io
        
        if not req.documents:
            df = pd.DataFrame(columns=req.columns)
        else:
            df = pd.DataFrame(req.documents)
        
        # Strip out metadata columns like __id or is_dirty if present
        cols_to_export = [col for col in req.columns if col in df.columns]
        
        # Reorder columns
        if cols_to_export:
            df = df[cols_to_export]
            
        # Create an in-memory bytes buffer for the Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Export to sheet named VDL
            df.to_excel(writer, index=False, sheet_name="VDL")
            
            workbook = writer.book
            worksheet = writer.sheets["VDL"]
            
            # Freeze header row
            worksheet.freeze_panes = "A2"
            
            from openpyxl.styles import PatternFill, Font, Border, Side
            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
            
            # Format header row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                
            # Auto-adjust column widths for premium feel & apply borders to all cells
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    # Apply border to all cells
                    if cell.row > 1:
                        cell.border = thin_border
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        output.seek(0)
        
        # Return as StreamingResponse with correct headers
        filename = f"VDL_Export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e:
        log_error_to_file("Export failed", e)
        raise HTTPException(status_code=500, detail=f"Errore durante l'esportazione in Excel: {str(e)}")

@app.post("/api/project/{project_id}/add-cycle")
def add_project_cycle(project_id: int, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    import re
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Progetto non trovato.")
    
    columns = project["columns"]
    # Find max cycle number
    max_cycle = 0
    for col in columns:
        norm_k = col.lower().replace(" ", "").strip()
        if norm_k.startswith("returncode"):
            match = re.search(r'returncode(\d*)', norm_k)
            if match:
                num_str = match.group(1)
                num = int(num_str) if num_str else 0
                max_cycle = max(max_cycle, num)
                
    next_cycle = max(1, max_cycle + 1)
    
    new_cols = [
        f"TR Out{next_cycle}",
        f"Actual Date{next_cycle}",
        f"TR In{next_cycle}",
        f"Return Code{next_cycle}",
        f"Return Date{next_cycle}"
    ]
    
    # Avoid duplicates and append
    for c in new_cols:
        if c not in columns:
            columns.append(c)
            
    # Update project columns safely
    db.update_project_columns(project_id, columns)
    
    # Update all documents adding the empty fields
    docs = db.get_documents(project_id)
    for doc in docs:
        doc_id = doc.pop("__id")
        for c in new_cols:
            if c not in doc:
                doc[c] = ""
        db.update_document(doc_id, doc)
        
    return {"status": "success", "message": f"Ciclo di revisione {next_cycle} aggiunto con successo.", "new_columns": new_cols}

@app.get("/api/project/{project_id}/status")
def get_project_status(project_id: int):
    """
    Returns a breakdown of document statuses based on Last Code receive and TR Out columns.
    Categories:
      - per_code: dict { code_value -> count }, for docs that have a Last Code receive value
      - pending: count of docs that have at least one TR Out filled but no Last Code receive
      - never_emitted: count of docs that have no TR Out filled at all (never sent)
    """
    import re
    try:
        project = db.get_project(project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Progetto non trovato.")

        docs = db.get_documents(project_id)
        settings = db.get_settings()
        exchange_time = int(settings.get("exchange_time", "15") or "15")
        for doc in docs:
            recalculate_computed_fields(doc, exchange_time)

        # Identify all TR Out columns (TR Out, TR Out1, TR Out2, ...)
        columns = project.get("columns", [])
        tr_out_cols = []
        for col in columns:
            norm = col.lower().replace(" ", "").strip()
            if re.fullmatch(r'trout\d*', norm):
                tr_out_cols.append(col)

        per_code: Dict[str, int] = {}
        pending_count = 0
        never_emitted_count = 0

        contractor_clean = (project.get("contractor") or "").strip().lower()

        for doc in docs:
            last_code = ""
            for k, v in doc.items():
                if k.lower().replace(" ", "").strip() == "lastcodereceive":
                    last_code = str(v).strip() if v is not None else ""
                    if last_code.lower() in ["nan", "nat", "null", "-"]:
                        last_code = ""
                    break

            hands_val = ""
            for k, v in doc.items():
                if k.lower().replace(" ", "").strip() == "hands":
                    hands_val = str(v).strip() if v is not None else ""
                    break

            is_hands_contractor = False
            if contractor_clean and hands_val.lower() == contractor_clean:
                is_hands_contractor = True

            # Check if any TR Out column has a non-empty value
            has_tr_out = False
            for col in tr_out_cols:
                val = doc.get(col, "")
                if val and str(val).strip() and str(val).strip().lower() not in ["nan", "nat", "null", "-", ""]:
                    has_tr_out = True
                    break

            if is_hands_contractor:
                # Document is with the Contractor (pending return)
                pending_count += 1
            elif last_code:
                # Document has been reviewed and returned with a code
                per_code[last_code] = per_code.get(last_code, 0) + 1
            elif has_tr_out:
                # Fallback: Document was sent (TR Out filled) but hasn't come back yet
                pending_count += 1
            else:
                # Document was never sent (no TR Out at all)
                never_emitted_count += 1

        return {
            "status": "success",
            "total": len(docs),
            "per_code": per_code,
            "pending": pending_count,
            "never_emitted": never_emitted_count
        }
    except HTTPException:
        raise
    except Exception as e:
        log_error_to_file("Project status calculation failed", e)
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/project/{project_id}/cross-warnings")
def get_cross_vdl_warnings(project_id: int):
    """
    Returns a list of warnings for documents that have been returned by the supplier
    but either do not exist in the client VDL or have not been sent to the client
    (or sent before the latest supplier update).
    """
    try:
        project = db.get_project(project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Progetto non trovato.")

        suppliers = db.get_project_suppliers(project_id)
        client_docs = db.get_documents(project_id)
        
        # We need to map client docs by their Internal Document Number
        client_map = {}
        for doc in client_docs:
            doc_num = get_doc_code(doc)
            if doc_num and doc_num != "-":
                client_map[doc_num] = doc
                
        warnings = []
        import re
        
        for supplier in suppliers:
            supp_docs = db.get_supplier_documents(project_id, supplier["id"])
            
            for s_doc in supp_docs:
                doc_num = get_doc_code(s_doc)
                if not doc_num or doc_num == "-":
                    continue # Ignore docs with no internal number
                # We want to compare Next issue forecast date
                s_forecast_str = get_row_val(s_doc, "Next issue forecast date")
                if not s_forecast_str:
                    continue # Supplier hasn't given a forecast
                s_forecast_dt = parse_date(s_forecast_str)
                if not s_forecast_dt:
                    continue
                    
                # Now check client VDL
                c_doc = client_map.get(doc_num)
                if not c_doc:
                    continue # Not in client VDL
                    
                c_forecast_str = get_row_val(c_doc, "Next issue forecast date")
                if not c_forecast_str:
                    continue # Client hasn't got a forecast
                c_forecast_dt = parse_date(c_forecast_str)
                if not c_forecast_dt:
                    continue
                    
                desc = get_doc_title(c_doc)
                
                # If Client expects it BEFORE Supplier says it will be ready
                if c_forecast_dt < s_forecast_dt:
                    warnings.append({
                        "supplier_name": supplier["name"],
                        "document_number": doc_num,
                        "description": str(desc),
                        "supplier_date": s_forecast_dt.strftime('%Y-%m-%d'),
                        "client_date": c_forecast_dt.strftime('%Y-%m-%d'),
                        "reason": f"Ritardo Fornitore! Previsto al Cliente: {c_forecast_dt.strftime('%d/%m/%Y')}. Promesso dal Fornitore: {s_forecast_dt.strftime('%d/%m/%Y')}."
                    })

        return {"status": "success", "warnings": warnings}
    except Exception as e:
        log_error_to_file("Cross warnings check failed", e)
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# SUPPLIER VDL ENDPOINTS
# ==========================================

@app.post("/api/supplier-vdl/{project_id}/{supplier_id}/import")
async def import_supplier_vdl(project_id: int, supplier_id: int, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    try:
        filename = file.filename
        content = await file.read()
        
        if filename.endswith(".csv"):
            import io
            df = pd.read_csv(io.BytesIO(content))
        elif filename.endswith((".xlsx", ".xls")):
            import io
            df = pd.read_excel(io.BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Formato file non supportato. Carica un file Excel o CSV.")
            
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime('%Y-%m-%d')
                
        raw_rows = df.to_dict(orient="records")
        cleaned_rows = []
        import datetime
        for row in raw_rows:
            clean_row = {}
            for k, v in row.items():
                if pd.isna(v):
                    clean_row[k] = ""
                elif isinstance(v, (datetime.datetime, datetime.date)):
                    clean_row[k] = v.strftime('%Y-%m-%d')
                else:
                    clean_row[k] = v
            cleaned_rows.append(clean_row)
            
        df = pd.DataFrame(cleaned_rows)
        existing_cols = list(df.columns)
        
        columns_to_add = []
        for custom_col in CUSTOM_TRACKING_COLUMNS:
            normalized_custom = custom_col.lower().replace(" ", "").strip()
            found = False
            for exist_col in existing_cols:
                if exist_col.lower().replace(" ", "").strip() == normalized_custom:
                    found = True
                    break
            if not found:
                columns_to_add.append(custom_col)
                
        for col in columns_to_add:
            df[col] = None
            
        final_columns = list(df.columns)
        rows = df.to_dict(orient="records")
        
        settings = db.get_settings()
        exchange_time = int(settings.get("exchange_time", "15") or "15")
        for row in rows:
            recalculate_computed_fields(row, exchange_time)
            
        # Update columns schema in project_suppliers mapping
        db.set_project_supplier_columns(project_id, supplier_id, final_columns)
        
        # Replace existing docs
        db.clear_supplier_documents(project_id, supplier_id)
        db.save_supplier_documents(project_id, supplier_id, rows)
        
        return {"status": "success", "message": f"VDL Fornitore importata ({len(rows)} righe)."}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/supplier-vdl/{project_id}/{supplier_id}/documents")
def get_supplier_vdl_docs(project_id: int, supplier_id: int, current_user: dict = Depends(get_current_user)):
    docs = db.get_supplier_documents(project_id, supplier_id)
    cols = db.get_project_supplier_columns(project_id, supplier_id)
    return {"status": "success", "columns": cols, "documents": recalculate_row_list(docs)}

@app.post("/api/supplier-vdl/{project_id}/{supplier_id}/save-all")
def save_supplier_vdl_all(project_id: int, supplier_id: int, documents: List[Dict[str, Any]], current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    settings = db.get_settings()
    exchange_time = int(settings.get("exchange_time", "15") or "15")
    
    for doc in documents:
        recalculate_computed_fields(doc, exchange_time)
        
    db.save_supplier_documents(project_id, supplier_id, documents)
    return {"status": "success"}

@app.post("/api/supplier-vdl/{project_id}/{supplier_id}/add-cycle")
def add_supplier_vdl_cycle(project_id: int, supplier_id: int, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    import re
    columns = db.get_project_supplier_columns(project_id, supplier_id)
    if not columns:
        raise HTTPException(status_code=400, detail="Schema non trovato.")
    
    max_cycle = 0
    for col in columns:
        norm_k = col.lower().replace(" ", "").strip()
        if norm_k.startswith("returncode"):
            match = re.search(r'returncode(\d*)', norm_k)
            if match:
                num_str = match.group(1)
                max_cycle = max(max_cycle, int(num_str) if num_str else 0)
                
    next_cycle = max(1, max_cycle + 1)
    
    new_cols = [
        f"TR Out{next_cycle}", f"Actual Date{next_cycle}",
        f"TR In{next_cycle}", f"Return Code{next_cycle}", f"Return Date{next_cycle}"
    ]
    
    for c in new_cols:
        if c not in columns:
            columns.append(c)
            
    db.update_supplier_project_columns(project_id, supplier_id, columns)
    return {"status": "success", "new_columns": new_cols}

@app.get("/api/supplier-vdl/{project_id}/{supplier_id}/export")
def export_supplier_vdl(project_id: int, supplier_id: int, current_user: dict = Depends(get_current_user)):
    try:
        docs = db.get_supplier_documents(project_id, supplier_id)
        cols = db.get_project_supplier_columns(project_id, supplier_id)
        if not docs or not cols:
            raise HTTPException(status_code=404, detail="Nessun dato da esportare.")
            
        settings = db.get_settings()
        exchange_time = int(settings.get("exchange_time", "15") or "15")
        for doc in docs:
            recalculate_computed_fields(doc, exchange_time)
            
        filtered = []
        for doc in docs:
            filtered_doc = {k: doc.get(k, "") for k in cols}
            filtered.append(filtered_doc)
            
        df = pd.DataFrame(filtered)
        
        import io
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="VDL Fornitore", index=False)
            
            worksheet = writer.sheets["VDL Fornitore"]
            worksheet.freeze_panes = "A2"
            
            from openpyxl.styles import PatternFill, Font, Border, Side
            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.row > 1:
                        cell.border = thin_border
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        output.seek(0)
        filename = f"VDL_Fornitore_{supplier_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}", "Access-Control-Expose-Headers": "Content-Disposition"}
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# Contacts Endpoints
@app.get("/api/contacts")
def get_contacts(project_id: int, current_user: dict = Depends(get_current_user)):
    return db.get_contacts(project_id)

@app.post("/api/contacts")
def save_contact(contact: ContactSave, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    db.save_contact(contact.project_id, contact.hands_value, contact.to_emails, contact.cc_emails)
    return {"status": "success", "message": "Contatto salvato."}

@app.delete("/api/contacts/{project_id}/{hands_value}")
def delete_contact(project_id: int, hands_value: str, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    db.delete_contact(project_id, hands_value)
    return {"status": "success", "message": "Contatto eliminato."}

@app.post("/api/contacts/upload")
async def upload_contacts(project_id: int = Form(...), file: UploadFile = File(...), current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    try:
        content = await file.read()
        import io
        if file.filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
            
        required = ["Nome fornitore", "PM fornitore", "PE fornitore", "PM interno", "PE interno", "DC fornitore"]
        for c in required:
            if c not in df.columns:
                raise HTTPException(status_code=400, detail=f"Colonna mancante nel file: {c}")
                
        count = 0
        for idx, row in df.iterrows():
            supplier = str(row.get("Nome fornitore", "")).strip()
            if not supplier or supplier.lower() == "nan": continue
            
            to_list = [str(row.get("PM fornitore", "")), str(row.get("PE fornitore", "")), str(row.get("DC fornitore", ""))]
            to_list = [t.strip() for t in to_list if t and str(t).lower() != "nan"]
            
            cc_list = [str(row.get("PM interno", "")), str(row.get("PE interno", ""))]
            cc_list = [c.strip() for c in cc_list if c and str(c).lower() != "nan"]
            
            db.save_contact(project_id, supplier, ", ".join(to_list), ", ".join(cc_list))
            count += 1
            
        return {"status": "success", "message": f"Rubrica aggiornata con {count} contatti."}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# Settings Endpoints
@app.get("/api/settings")
def get_settings(current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller", "Project Engineering"]))):
    return db.get_settings()

@app.post("/api/settings")
def save_settings(settings: SettingsSave, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller", "Project Engineering"]))):
    db.save_setting("sender_email", settings.sender_email)
    db.save_setting("smtp_server", settings.smtp_server)
    db.save_setting("smtp_port", settings.smtp_port)
    db.save_setting("smtp_user", settings.smtp_user)
    db.save_setting("smtp_password", settings.smtp_password)
    db.save_setting("email_mode", settings.email_mode)
    db.save_setting("exchange_time", settings.exchange_time)
    return {"status": "success", "message": "Impostazioni salvate."}


# --- SUPPLIERS ENDPOINTS ---

@app.get("/api/suppliers")
def get_all_suppliers():
    try:
        suppliers = db.get_all_suppliers()
        return {"status": "success", "suppliers": suppliers}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/suppliers")
def create_new_supplier(supplier: SupplierCreate):
    try:
        sid = db.create_supplier(supplier.name, supplier.item)
        return {"status": "success", "message": "Fornitore creato con successo.", "id": sid}
    except Exception as e:
        traceback.print_exc()
        if "UNIQUE" in str(e).upper():
            raise HTTPException(status_code=400, detail="Un fornitore con questo nome esiste già.")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/suppliers/{supplier_id}")
def update_existing_supplier(supplier_id: int, supplier: SupplierUpdate):
    try:
        db.update_supplier(supplier_id, supplier.name, supplier.item)
        return {"status": "success", "message": "Fornitore aggiornato con successo."}
    except Exception as e:
        traceback.print_exc()
        if "UNIQUE" in str(e).upper():
            raise HTTPException(status_code=400, detail="Un fornitore con questo nome esiste già.")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/suppliers/{supplier_id}")
def delete_existing_supplier(supplier_id: int):
    try:
        db.delete_supplier(supplier_id)
        return {"status": "success", "message": "Fornitore eliminato con successo."}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# --- PROJECT SUPPLIERS ASSIGNMENTS ENDPOINTS ---

@app.get("/api/project/{project_id}/suppliers")
def get_project_suppliers(project_id: int):
    try:
        suppliers = db.get_project_suppliers(project_id)
        return {"status": "success", "suppliers": suppliers}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/project/{project_id}/suppliers")
def sync_project_suppliers(project_id: int, payload: ProjectSuppliersSync):
    try:
        db.save_project_suppliers(project_id, payload.supplier_ids)
        return {"status": "success", "message": "Fornitori assegnati al progetto con successo."}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# --- SUPPLIER TRANSMITTALS ENDPOINTS ---

@app.get("/api/project/{project_id}/supplier-transmittals")
def get_supplier_transmittals(project_id: int):
    try:
        transmittals = db.get_supplier_transmittals(project_id)
        return {"status": "success", "transmittals": transmittals}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/supplier-transmittals")
def create_supplier_transmittal(req: SupplierTransmittalCreate):
    try:
        tr_id = db.create_supplier_transmittal(
            req.project_id,
            req.supplier_id,
            req.direction,
            req.tr_number,
            req.tr_date,
            req.notes or "",
            req.document_list
        )
        return {
            "status": "success",
            "message": f"Transmittal Fornitore {req.direction} registrato con successo.",
            "id": tr_id
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/supplier-transmittals/{tr_id}")
def delete_supplier_transmittal(tr_id: int):
    try:
        db.delete_supplier_transmittal(tr_id)
        return {"status": "success", "message": "Transmittal Fornitore eliminato."}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# Core Logic: Overdue Calculation
def calculate_delay_for_row(row: Dict[str, Any], exchange_time_days: int) -> Dict[str, Any]:
    """
    Checks if a document is overdue for issue by:
      - Finding the next expected Actual Date column in the lifecycle.
      - Checking if that column is empty (not issued yet).
      - Comparing today's date against 'Next issue forecast date' (the deadline).
    """
    today = datetime.date.today()
    
    # 1. Get Next issue forecast date
    forecast_val = get_row_val(row, "Next issue forecast date")
    next_forecast = parse_date(forecast_val) if forecast_val else None
    
    # 2. Map of Return Date columns to their corresponding next Actual Date columns
    return_cols = [
        ("Return Date6", None),
        ("Return Date5", "Actual Date6"),
        ("Return Date4", "Actual Date5"),
        ("Return Date3", "Actual Date4"),
        ("Return Date2", "Actual Date3"),
        ("Return Date1", "Actual Date2"),
        ("Return Date", "Actual Date1")
    ]
    
    waiting_for_issue = False
    next_actual_col = None
    last_filled_col = None
    last_date_str = ""
    
    # Scan Return Dates from latest (6) to earliest (0)
    found_return = False
    for ret_col, next_act_col in return_cols:
        ret_val = get_row_val(row, ret_col)
        if ret_val:
            dt = parse_date(ret_val)
            if dt:
                found_return = True
                last_filled_col = ret_col
                last_date_str = dt.strftime("%Y-%m-%d")
                
                if next_act_col:
                    next_act_col_val = get_row_val(row, next_act_col)
                    act_dt = parse_date(next_act_col_val) if next_act_col_val else None
                    if act_dt is None:
                        waiting_for_issue = True
                        next_actual_col = next_act_col
                    else:
                        waiting_for_issue = False
                        next_actual_col = next_act_col
                else:
                    # Return Date6 is filled, there is no next revision
                    waiting_for_issue = False
                break
                
    if not found_return:
        # No Return Date is filled at all. We check the first submission (Actual Date).
        first_act_val = get_row_val(row, "Actual Date")
        first_act_dt = parse_date(first_act_val) if first_act_val else None
        if first_act_dt is None:
            # We are waiting for the very first submission
            waiting_for_issue = True
            next_actual_col = "Actual Date"
            
            # Find the baseline date for user preview
            promise_val = get_row_val(row, "Promise date")
            promise_dt = parse_date(promise_val) if promise_val else None
            if promise_dt:
                last_filled_col = "Promise date"
                last_date_str = promise_dt.strftime("%Y-%m-%d")
            else:
                prev_val = get_row_val(row, "1° Invio previsione")
                prev_dt = parse_date(prev_val) if prev_val else None
                if prev_dt:
                    last_filled_col = "1° Invio previsione"
                    last_date_str = prev_dt.strftime("%Y-%m-%d")
        else:
            # First submission has been done
            waiting_for_issue = False
            last_filled_col = "Actual Date"
            last_date_str = first_act_dt.strftime("%Y-%m-%d")
            
    # Determine if overdue
    is_overdue = False
    delay_days = 0
    
    if waiting_for_issue and next_forecast:
        days_passed = (today - next_forecast).days
        if days_passed > 0:
            delay_days = days_passed
            if delay_days > exchange_time_days:
                is_overdue = True
            
    return {
        "has_date": next_forecast is not None,
        "last_filled_column": last_filled_col or "N/A",
        "last_filled_date": last_date_str or None,
        "next_issue_forecast_date": next_forecast.strftime("%Y-%m-%d") if next_forecast else None,
        "is_overdue": is_overdue,
        "delay_days": delay_days
    }


@app.post("/api/reminder/calculate")
def calculate_reminders(req: ReminderCalculateRequest):
    project = db.get_project(req.project_id)
    if not project:
        return {"status": "error", "message": "Progetto non trovato."}
        
    # Get global exchange time setting once as the source of truth
    settings = db.get_settings()
    exchange_time_days = int(settings.get("exchange_time", "15") or "15")
    
    documents = db.get_documents(project["id"])
    overdue_docs = []
    unique_hands_values = set()
    
    # First extract unique hands values
    for doc in documents:
        # Flexibly find Hands column
        hands_key = None
        for key in doc.keys():
            if key.lower().strip() == "hands":
                hands_key = key
                break
        if hands_key and doc[hands_key]:
            val = str(doc[hands_key]).strip()
            if val:
                unique_hands_values.add(val)
                
    # Calculate delay details
    for doc in documents:
        # Find Hands value
        hands_key = None
        hands_val = ""
        for key in doc.keys():
            if key.lower().strip() == "hands":
                hands_key = key
                break
        if hands_key and doc[hands_key]:
            hands_val = str(doc[hands_key]).strip()
            
        # If filtering by specific Hands, skip others
        if req.hands_value and req.hands_value != "all" and hands_val != req.hands_value:
            continue
            
        # Skip if hands is empty and we are filtering
        if req.hands_value and req.hands_value != "all" and not hands_val:
            continue
            
        # Calculate delay info using stored next issue forecast date and the global exchange time setting
        delay_info = calculate_delay_for_row(doc, exchange_time_days)
        if delay_info["is_overdue"]:
            doc_id_code = get_doc_code(doc)
            doc_title = get_doc_title(doc)
                    
            overdue_docs.append({
                "__id": doc.get("__id"),
                "document_code": doc_id_code,
                "document_title": doc_title,
                "hands": hands_val or "Non specificato",
                "last_filled_column": delay_info["last_filled_column"],
                "last_filled_date": delay_info["last_filled_date"],
                "next_issue_forecast_date": delay_info["next_issue_forecast_date"],
                "delay_days": delay_info["delay_days"],
                "raw_data": doc
            })
            
    return {
        "status": "success",
        "overdue_documents": overdue_docs,
        "unique_hands": sorted(list(unique_hands_values))
    }

# Email Generation & Sending
def generate_html_email(hands_value: str, overdue_docs: List[Dict[str, Any]], exchange_time: int, additional_notes: str, language: str = "it") -> str:
    today_str = datetime.date.today().strftime("%d/%m/%Y")
    
    is_en = (language.lower() == "en")
    
    # Translations
    t_title = "VDL Documents Delivery Reminder" if is_en else "Sollecito Consegna Documenti VDL"
    t_subtitle = f"Overdue documents report for: <strong>{hands_value}</strong>" if is_en else f"Report dei documenti in ritardo per l'attore: <strong>{hands_value}</strong>"
    t_greeting = "Dear " + hands_value + ",<br><br>Please find attached/summarized the status of the Vendor Document List (VDL) documents assigned to you which are currently overdue for issue or comment as of <strong>{today_str}</strong>." if is_en else f"Gentile {hands_value}, <br><br>Si trasmette in allegato/riassunto lo stato dei documenti Vendor Document List (VDL) a voi assegnati che risultano attualmente in ritardo di emissione o di commento alla data odierna del <strong>{today_str}</strong>."
    
    t_exchange = f"The expected document exchange time is set to <strong>{exchange_time} days</strong> from the last recorded transmission." if is_en else f"Il tempo previsto per lo scambio documentale è impostato a <strong>{exchange_time} giorni</strong> dall'ultimo invio o ricezione registrati."
    
    t_notes_title = "Additional Notes:" if is_en else "Note Aggiuntive:"
    t_table_title = "List of Overdue Documents" if is_en else "Lista dei Documenti in Ritardo"
    
    t_col1 = "Doc. Code" if is_en else "Codice Doc."
    t_col2 = "Title / Description" if is_en else "Titolo / Descrizione"
    t_col3 = "Doc Status" if is_en else "Stato Doc"
    t_col4 = "Last Date" if is_en else "Data Ultimo"
    t_col6 = "Delay" if is_en else "Ritardo"
    
    t_action = "<strong>Required Action:</strong> Please check the progress of the above documents with priority and proceed with the upload or submission of the relevant comments as soon as possible." if is_en else "<strong>Azione Richiesta:</strong> Si prega di verificare con priorità lo stato di avanzamento dei suddetti documenti e procedere con l'upload o con l'invio dei relativi commenti nel più breve tempo possibile."
    
    # Rows for the table
    table_rows = ""
    for idx, doc in enumerate(overdue_docs):
        row_bg = "#ffffff" if idx % 2 == 0 else "#f8fafc"
        delay_txt = f"{doc['delay_days']} days" if is_en else f"{doc['delay_days']} gg"
        
        # Get Returned Code status
        last_code = get_row_val(doc.get('raw_data', {}), 'Last Code receive')
        last_code_str = str(last_code).strip() if last_code is not None else ""
        if last_code_str.lower() in ["nan", "nat", "null", "-"]:
            last_code_str = ""
        doc_status = last_code_str if last_code_str else "-"

        table_rows += f"""
        <tr style="background-color: {row_bg};">
            <td style="padding: 10px; border: 1px solid #cbd5e1; font-family: Arial, sans-serif; font-size: 13px; color: #1e293b; font-weight: bold;">{doc['document_code']}</td>
            <td style="padding: 10px; border: 1px solid #cbd5e1; font-family: Arial, sans-serif; font-size: 13px; color: #475569;">{doc['document_title']}</td>
            <td style="padding: 10px; border: 1px solid #cbd5e1; font-family: Arial, sans-serif; font-size: 13px; color: #475569; font-style: italic;">{doc_status}</td>
            <td style="padding: 10px; border: 1px solid #cbd5e1; font-family: Arial, sans-serif; font-size: 13px; color: #475569;">{doc['last_filled_date']}</td>
            <td style="padding: 10px; border: 1px solid #cbd5e1; font-family: Arial, sans-serif; font-size: 13px; color: #e11d48; font-weight: bold; text-align: center;">{delay_txt}</td>
        </tr>
        """

    notes_section = ""
    if additional_notes:
        notes_section = f"""
        <div style="margin-bottom: 24px;">
            <p style="margin: 0 0 4px 0; color: #0369a1; font-family: Arial, sans-serif; font-weight: bold;">{t_notes_title}</p>
            <p style="margin: 0; color: #0c4a6e; font-family: Arial, sans-serif; font-size: 14px; line-height: 1.5;">{additional_notes}</p>
        </div>
        """

    t_greeting_formatted = t_greeting.format(today_str=today_str)
    
    # Simplified Outlook-compatible HTML structure without signature
    html = f"""
        <div style="max-width: 800px; margin: 0 auto; background-color: #ffffff; font-family: Arial, sans-serif; color: #334155;">
            <!-- Header -->
            <div style="background-color: #1e3a8a; padding: 24px; text-align: left; color: #ffffff;">
                <h2 style="margin: 0; font-family: Arial, sans-serif; font-weight: bold;">{t_title}</h2>
                <p style="margin: 8px 0 0 0; font-size: 14px; font-family: Arial, sans-serif;">{t_subtitle}</p>
            </div>
            
            <!-- Content -->
            <div style="padding: 24px; font-family: Arial, sans-serif; font-size: 14px; line-height: 1.5;">
                <p style="margin-top: 0;">{t_greeting_formatted}</p>
                <p style="margin-bottom: 24px;">{t_exchange}</p>
                
                {notes_section}
                
                <h3 style="color: #1e3a8a; margin-top: 32px; font-family: Arial, sans-serif;">{t_table_title}</h3>
                
                <table style="width: 100%; border-collapse: collapse; text-align: left; margin-bottom: 32px; border: 1px solid #cbd5e1;">
                    <thead>
                        <tr style="background-color: #f1f5f9;">
                            <th style="padding: 10px; border: 1px solid #cbd5e1; font-family: Arial, sans-serif; font-size: 13px; font-weight: bold; color: #475569; width: 25%;">{t_col1}</th>
                            <th style="padding: 10px; border: 1px solid #cbd5e1; font-family: Arial, sans-serif; font-size: 13px; font-weight: bold; color: #475569; width: 40%;">{t_col2}</th>
                            <th style="padding: 10px; border: 1px solid #cbd5e1; font-family: Arial, sans-serif; font-size: 13px; font-weight: bold; color: #475569; width: 15%;">{t_col3}</th>
                            <th style="padding: 10px; border: 1px solid #cbd5e1; font-family: Arial, sans-serif; font-size: 13px; font-weight: bold; color: #475569; width: 10%;">{t_col4}</th>
                            <th style="padding: 10px; border: 1px solid #cbd5e1; font-family: Arial, sans-serif; font-size: 13px; font-weight: bold; color: #475569; width: 10%; text-align: center;">{t_col6}</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows}
                    </tbody>
                </table>
                
                <p style="background-color: #f8fafc; padding: 12px; border: 1px solid #cbd5e1;">{t_action}</p>
                <br><br>
            </div>
        </div>
    """
    return html

@app.post("/api/reminder/preview")
def preview_reminder(req: EmailSendRequest, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    project = db.get_project(req.project_id)
    if not project:
        raise HTTPException(status_code=400, detail="Progetto non trovato.")
        
    calc_res = calculate_reminders(ReminderCalculateRequest(project_id=req.project_id, hands_value=req.hands_value, exchange_time=req.exchange_time, language=req.language))
    overdue_docs = calc_res.get("overdue_documents", [])
    
    if not overdue_docs:
        return {
            "status": "warning", 
            "message": f"Nessun documento in ritardo trovato per '{req.hands_value}'.",
            "html": ""
        }
        
    contacts = db.get_contacts(req.project_id)
    contact = contacts.get(req.hands_value, {"to": "", "cc": ""})
    
    html_content = generate_html_email(req.hands_value, overdue_docs, req.exchange_time, req.additional_notes or "", req.language)
    
    return {
        "status": "success",
        "to": str(contact.get("to") or ""),
        "cc": str(contact.get("cc") or ""),
        "subject": req.subject,
        "html": html_content,
        "count": len(overdue_docs)
    }

@app.post("/api/reminder/send")
def send_reminder(req: EmailSendRequest, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    try:
        return _send_reminder_impl(req)
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Errore interno: {str(e)}")

def merge_html_and_signature(vdl_html: str, signature_html: str) -> str:
    """
    Merges VDL reminder HTML content into the Outlook signature HTML
    to form a single, well-formed HTML document.
    """
    if not signature_html:
        return vdl_html
        
    import re
    # Case-insensitive search for <body...> tag
    match = re.search(r'<body[^>]*>', signature_html, re.IGNORECASE)
    if match:
        body_end_idx = match.end()
        # Insert vdl_html right after the opening <body> tag
        merged = signature_html[:body_end_idx] + vdl_html + signature_html[body_end_idx:]
        return merged
        
    # If no <body> tag is found, fallback to prepending vdl_html to signature_html
    return vdl_html + signature_html

def get_outlook_signature_html(sender_email: str) -> str:
    """
    Reads the signature of the selected sender email directly from AppData/Microsoft/Signatures
    and replaces relative image paths with absolute file:// paths so they display correctly.
    """
    try:
        import os
        import urllib.parse
        appdata = os.environ.get("APPDATA", "")
        if not appdata:
            return ""
        sig_dir = os.path.join(appdata, "Microsoft", "Signatures")
        if not os.path.exists(sig_dir):
            return ""
            
        sig_files = [f for f in os.listdir(sig_dir) if f.endswith(".htm")]
        if not sig_files:
            return ""
            
        chosen_file = None
        sender_lower = sender_email.lower() if sender_email else ""
        
        if sender_lower:
            # 1. Match by containing parts (exact email match first)
            for f in sig_files:
                if sender_lower in f.lower():
                    chosen_file = f
                    break
            
            # 2. Match by domain/name keywords
            if not chosen_file:
                if "emtb" in sender_lower:
                    for f in sig_files:
                        if "emtb" in f.lower():
                            chosen_file = f
                            break
                elif "francotosimeccanica" in sender_lower or "tosi" in sender_lower:
                    for f in sig_files:
                        if "ftm" in f.lower() or "tosi" in f.lower() or "francotosimeccanica" in f.lower():
                            chosen_file = f
                            break
                elif "satiz" in sender_lower or "satiztpm" in sender_lower:
                    for f in sig_files:
                        if "satiz" in f.lower() or "satiztpm" in f.lower():
                            chosen_file = f
                            break
                            
        if not chosen_file:
            # Fallback to the first signature file if no match
            chosen_file = sig_files[0]
            
        if chosen_file:
            filepath = os.path.join(sig_dir, chosen_file)
            with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
                
            sig_name_no_ext = os.path.splitext(chosen_file)[0]
            encoded_sig_name = urllib.parse.quote(sig_name_no_ext)
            
            folder_name = sig_name_no_ext + "_files"
            folder_name_alt = sig_name_no_ext + "_file"
            
            encoded_folder = encoded_sig_name + "_files"
            encoded_folder_alt = encoded_sig_name + "_file"
            
            abs_folder_path = "file:///" + os.path.join(sig_dir, folder_name).replace("\\", "/")
            abs_folder_path_alt = "file:///" + os.path.join(sig_dir, folder_name_alt).replace("\\", "/")
            
            # Replace forward slashes
            content = content.replace(folder_name + "/", abs_folder_path + "/")
            content = content.replace(folder_name_alt + "/", abs_folder_path_alt + "/")
            content = content.replace(encoded_folder + "/", abs_folder_path + "/")
            content = content.replace(encoded_folder_alt + "/", abs_folder_path_alt + "/")
            
            # Replace backslashes
            content = content.replace(folder_name + "\\", abs_folder_path + "/")
            content = content.replace(folder_name_alt + "\\", abs_folder_path_alt + "/")
            content = content.replace(encoded_folder + "\\", abs_folder_path + "/")
            content = content.replace(encoded_folder_alt + "\\", abs_folder_path_alt + "/")
            
            return content
    except Exception as e:
        print(f"Error loading signature file: {e}")
    return ""

def _send_reminder_impl(req: EmailSendRequest):
    project = db.get_project(req.project_id)
    if not project:
        raise HTTPException(status_code=400, detail="Progetto non trovato.")
        
    settings = db.get_settings()
    email_mode = settings.get("email_mode", "outlook").lower()
    
    calc_res = calculate_reminders(ReminderCalculateRequest(project_id=req.project_id, hands_value=req.hands_value, exchange_time=req.exchange_time, language=req.language))
    overdue_docs = calc_res.get("overdue_documents", [])
    
    if not overdue_docs:
        raise HTTPException(status_code=400, detail=f"Nessun documento in ritardo per '{req.hands_value}' da sollecitare.")
        
    contacts = db.get_contacts(req.project_id)
    contact = contacts.get(req.hands_value, {"to": "", "cc": ""})
    
    to_emails = str(contact.get("to") or "").strip()
    cc_emails = str(contact.get("cc") or "").strip()
    
    if not to_emails:
        raise HTTPException(status_code=400, detail=f"Nessun destinatario (To) configurato nella rubrica per '{req.hands_value}'.")
        
    html_content = generate_html_email(req.hands_value, overdue_docs, req.exchange_time, req.additional_notes, req.language)
    
    # Sending implementation
    if email_mode == "outlook":
        if sys.platform != "win32":
            # Graceful fallback on non-Windows
            email_mode = "mock"
        else:
            try:
                import win32com.client
                import pythoncom
                
                # Critical for running win32com in FastAPI threads!
                pythoncom.CoInitialize()
                
                try:
                    outlook = win32com.client.Dispatch("Outlook.Application")
                    mail = outlook.CreateItem(0)  # 0 = olMailItem
                    
                    # 1. Set specific sender account BEFORE Display so Outlook loads the correct signature!
                    if req.sender_email:
                        try:
                            mail.SentOnBehalfOfName = req.sender_email
                            for account in outlook.Session.Accounts:
                                if account.SmtpAddress.lower() == req.sender_email.lower():
                                    mail.SendUsingAccount = account
                                    print(f"Outlook COM: Set SendUsingAccount to {account.SmtpAddress}")
                                    break
                            print(f"Outlook COM: Set SentOnBehalfOfName to {req.sender_email}")
                        except Exception as e:
                            print("Error setting sender account:", e)

                    # 2. Display the window to trigger signature injection
                    mail.Display()
                    
                    # 3. Retrieve default signature immediately
                    signature = mail.HTMLBody or ""
                    
                    # 4. Set recipient details and subject
                    mail.To = to_emails.replace(",", ";")
                    mail.CC = cc_emails.replace(",", ";")
                    mail.Subject = req.subject
                    
                    # 5. Prepend reminder content directly to signature
                    mail.HTMLBody = html_content + signature
                    
                    return {
                        "status": "success",
                        "method": "outlook",
                        "message": "Bozza aperta in Outlook. Controlla la finestra di Outlook sul tuo PC."
                    }
                except Exception as ex:
                    traceback.print_exc()
                    print("Outlook COM error:", ex)
                    log_error_to_file("Errore durante la preparazione dell'email con Outlook COM (in _send_reminder_impl)", ex)
                    # Fallback: return warning instead of crashing
                    return {
                        "status": "warning",
                        "method": "outlook_failed",
                        "message": f"Outlook non raggiungibile: {str(ex)}. Assicurati che Outlook sia aperto."
                    }
                finally:
                    pythoncom.CoUninitialize()
            except Exception as e:
                traceback.print_exc()
                log_error_to_file("Impossibile avviare Outlook COM (CoInitialize o dispatch fallito)", e)
                return {
                    "status": "warning",
                    "method": "outlook_failed",
                    "message": f"Errore avvio Outlook: {str(e)}"
                }
                
    if email_mode == "smtp":
        smtp_server = settings.get("smtp_server", "")
        smtp_port = int(settings.get("smtp_port", "587") or "587")
        smtp_user = settings.get("smtp_user", "")
        smtp_pass = settings.get("smtp_password", "")
        sender_email = req.sender_email or settings.get("sender_email", "")
        
        if not smtp_server or not sender_email:
            raise HTTPException(status_code=400, detail="Configurazione SMTP incompleta. Configura mittente e server SMTP nelle Impostazioni.")
            
        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = req.subject
            msg["From"] = sender_email
            msg["To"] = to_emails
            if cc_emails:
                msg["Cc"] = cc_emails
                
            part = MIMEText(html_content, "html")
            msg.attach(part)
            
            # Connect to SMTP
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.ehlo()
            
            # Start TLS if port is standard TLS/submission port
            if smtp_port == 587 or smtp_port == 25:
                server.starttls()
                server.ehlo()
                
            if smtp_user and smtp_pass:
                server.login(smtp_user, smtp_pass)
                
            # Send
            destinations = [email.strip() for email in to_emails.split(",") if email.strip()]
            if cc_emails:
                destinations += [email.strip() for email in cc_emails.split(",") if email.strip()]
                
            server.sendmail(sender_email, destinations, msg.as_string())
            server.quit()
            
            return {
                "status": "success",
                "method": "smtp",
                "message": f"Email di sollecito inviata con successo via SMTP a: {to_emails}!"
            }
        except Exception as e:
            traceback.print_exc()
            log_error_to_file("Errore durante l'invio SMTP (in _send_reminder_impl)", e)
            raise HTTPException(status_code=500, detail=f"Errore durante l'invio SMTP: {str(e)}")
            
    # Mock fallback (saving to local file)
    try:
        mock_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mock_emails")
        os.makedirs(mock_dir, exist_ok=True)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reminder_{req.hands_value}_{timestamp}.html"
        filepath = os.path.join(mock_dir, filename)
        
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html_content)
            
        return {
            "status": "success",
            "method": "mock",
            "message": f"Simulazione Invio: Email generata e salvata localmente in '{filename}'. Destinatari: A: {to_emails} | CC: {cc_emails}",
            "filepath": filepath
        }
    except Exception as e:
        log_error_to_file("Errore durante la scrittura del mock dell'email", e)
        raise HTTPException(status_code=500, detail=f"Errore nella generazione mock dell'email: {str(e)}")

# Define root route to serve index.html
static_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
if not os.path.exists(static_path):
    os.makedirs(static_path)
    os.makedirs(os.path.join(static_path, "css"))
    os.makedirs(os.path.join(static_path, "js"))

@app.get("/", response_class=HTMLResponse)
async def read_index():
    return FileResponse(os.path.join(static_path, "index.html"))

# Mount static files under /static
app.mount("/static", StaticFiles(directory=static_path), name="static")

if __name__ == "__main__":
    import uvicorn
    # Avvia su tutte le interfacce di rete (0.0.0.0) per renderlo accessibile ovunque e in qualsiasi rete
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)


@app.delete("/api/project/{project_id}")
def delete_project(project_id: int, current_user: dict = Depends(require_roles(["Admin", "Project Manager", "Document Controller"]))):
    try:
        db.delete_project(project_id)
        return {"status": "success", "message": "Progetto eliminato con successo"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
